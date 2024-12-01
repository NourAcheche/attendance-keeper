import tkinter
from tkinter import*
from tkinter import ttk
import openpyxl
class student:
    def __init__(self,first_name,last_name,section):
        self.first_name=first_name
        self.lastname=last_name
        self.section=section

#udemy
class attendance(Frame):
    def __init__(self, parent):
        Frame.__init__(self, parent)
        self.initUI()
        self.students=[]
        self.combo1.bind("<<ComboboxSelected>>",self.selected_list)
    def initUI(self):
        l = Label(self, text='AttendanceKeeper V1.0', font=("Arial", 12, "bold"))
        l.grid(row=0,column=0,columnspan=5)
        l1=Label(self,text="select student list Exel file: ")
        l1.grid(row=1,column=0,columnspan=1,sticky='E')
        l2=Label(self,text="select a student:" )
        l2.grid(row =2,column=0,columnspan=1)
        l3=Label(self,text="Section:")
        l3.grid(row=2,column=2)
        l4=Label(self,text="Attended students:")
        l4.grid(row=2,column=3,columnspan=4)
        l5 = Label(self, text="Please select file type:",font=("Arial",8))
        l5.grid(row=6, column=0, sticky='W')
        l6 = Label(self, text="Please enter week :")
        l6.grid(row=6, column=2, sticky='WENS')
        b1 = Button(self, text="import list",command=self.load_data)
        b1.grid(row=1, column=2)
        students = []
        attended_student=[]
        self.lb = Listbox(self, selectmode="multiple",height=5,width=25)
        for i in students:
            self.lb.insert(END, i)
        self.lb.grid(row=3,column=0,rowspan=3,columnspan=2,sticky='WENS')
        self.lb1=Listbox(self,selectmode="multiple",height=5,width=25)
        for i in attended_student:
            self.lb1.insert(END,i)
        self.lb1.grid(row=3,rowspan=3,column=3,columnspan=2,sticky='WENS')
        b2 = Button(self, text="Add=>",command=self.Add_student)
        b2.grid(row=4, column=2,sticky='WES')
        b3 = Button(self, text="<=Remove",command=self.delete_student)
        b3.grid(row=5, column=2,sticky='WES')
        self.input=Entry(self,width=12)
        self.input.grid(row=6,column=3)
        b4=Button(self,text="export as File",command=self.export_file)
        b4.grid(row=6, column=4, sticky='WES')
        box_values = [f"AP{i:02d}" for i in range(1, 21)]
        self.combo1 = ttk.Combobox(self, values=box_values)
        self.combo1.grid(row=3, column=2, padx=5, pady=5)
        self.combo1.set('AP01')
        self.combo2 = ttk.Combobox(self, values=[".xls", ".csv", ".txt"], state="readonly", width=3)
        self.combo2.grid(row=6, column=1, padx=5, pady=5, sticky='E')
        self.combo2.set('.txt')
#https://www.w3schools.com/python/ref_string_startswith.asp#:~:text=Python%20String%20startswith()%20Method
    def selected_list(self,e):
        selected_candidates=self.combo1.get()
        Nlist=[student for student in self.students if student[1].startswith(selected_candidates)]
        self.lb1.delete(0,tkinter.END)
        for i in Nlist:
            self.lb1.insert(tkinter.END,i)
    #https://youtu.be/8m4uDS_nyCk?si=R4hinfvdcti2KPOo

    def load_data(self):
        path = r"C:\Users\ASUS EXPERTBOOK I5\Downloads\AP - 2324 Spring - MP1 (2) (1)\Sample Input Student List\AP Student List.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        list_values=list(sheet.values)
        print(list_values)
        for row in sheet.iter_rows(values_only=True):
            student_name = row[1] ,',', str(row[0])
            self.students.append(student_name)
        self.lb.delete(0,tkinter.END)
        for i in self.students:
            self.lb.insert(tkinter.END,i)

    def delete_student(self):
        SI=self.lb1.curselection()
        for i in SI:
            self.lb1.delete(i)
    #https://www.w3schools.com/python/ref_list_pop.asp#:~:text=Python%20List%20pop()%20Method

    def Add_student(self):
        SI=self.lb1.curselection()
        for i in SI:
            student_name=self.lb1.get(i).split(',')
            self.lb1.insert(tkinter.END,student_name)
            self.students.append(self.students.pop(i))
    def export_file(self):
        attended_students=list(self.lb1.get(0,tkinter.END))
        file_type=self.combo2.get()
        week=self.input.get()

        if file_type=='.csv':
            raise BaseException("file type is not suppoerted ")
        elif file_type==".xls":
            workbook=openpyxl.Workbook()
            sheet=workbook.active
            sheet.append(["Id","Name","Dep"])
            for rec in self.students:
                sheet.append(rec[0],rec[1],rec[2])
        elif file_type==".txt":
            with open (f"{week}_attended_students.txt","a+") as file:
                for rec in attended_students:
                    line=f"{rec}"
                    file.write(line)






def main():
    root = Tk()
    root.title("tk")
    root.geometry("510x250")
    app = attendance(root)
    app.grid()
    root.mainloop()


main()
